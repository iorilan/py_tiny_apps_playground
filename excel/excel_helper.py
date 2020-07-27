import openpyxl, pprint
from openpyxl.styles import Font, NamedStyle, Color
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.layout import Layout, ManualLayout

"""
    {"sheet1":grid[][], "sheet2":grid[][]...}
"""
class ex_doc():
    def __init__(self):
        self.sheets = []

class excel_helper:
    def readSheet(self, file_name, sheet_name):
        print(f'opening {file_name} ...')
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.get_sheet_by_name(sheet_name)
        #print(sheet)
        cols = list(sheet.columns)
        #print(cols)
        data = []
        for r in range(1, sheet.max_row+1):
            row =[]
            for c in range(1, len(cols)+1):
                row.append(sheet.cell(row=r, column=c).value)
            data.append(row)
        wb.close()
        return data

    def readAll(self,file_name):
        print(f'opening {file_name} ...')
        wb = openpyxl.load_workbook(file_name)
        names= wb.get_sheet_names();
        res={}
        for sheet_name in names:
            #print(sheet_name)
            sheet = wb.get_sheet_by_name(sheet_name)
            #print(sheet)
            cols = list(sheet.columns)
            #print(cols)
            data = []
            for r in range(1, sheet.max_row+1):
                row =[]
                for c in range(1, len(cols)+1):
                    row.append(sheet.cell(row=r, column=c).value)
                data.append(row)
            res[sheet_name]=data
        wb.close()
        return res

    def write_excel_doc(self, data, file_name):
        wb = openpyxl.Workbook()
        wb.remove_sheet(wb.get_sheet_by_name("Sheet")) # remove default

        #https://openpyxl.readthedocs.io/en/stable/styles.html
        font1 = Font(name='Times New Roman', bold=True, size=24, italic=True)
        

        for sheet, grid in data.items():
            new_sheet = wb.create_sheet(sheet)
            for r in range(1, len(grid)+1):
                for c in range(1,len(grid[r-1])+1):
                    #print(r,c)
                    new_sheet.cell(row=r,column = c).value = grid[r-1][c-1]
                    new_sheet.cell(row=r,column=c).font = font1
        wb.save(file_name)
    
    
    def merge_with_sum_sample(self):
        marks = {
            'david':[100,99,98],
            'leo':[89,78,69.5],
            'jim':[91.5,59,67]
        }
        wb=openpyxl.Workbook()
        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        font1 = Font(name="Times New Roman", bold=True, size=22)
        font2 = Font(name="Times New Roman", bold=True, size=22,color='FF0000')
        font3 = Font(name="Times New Roman", bold=True, size=22,italic=True)

        sheet = wb.create_sheet("Marks")
        sheet.cell(row=1,column=1).value="Name"
        sheet.cell(row=1,column=2).value="Marks"
        sheet.merge_cells(f'B1:D1')

        """
        sheet.freeze_panes = 'A2'           Row 1
        sheet.freeze_panes = 'B1'           Column A
        sheet.freeze_panes = 'C1'           Columns A and B
        sheet.freeze_panes = 'C2'           Row 1 and columns A and B
        sheet.freeze_panes = 'A1'|None      no freeze
        """
        sheet.freeze_panes = "A2" # freeze row 1


        row = 2
        
        for name in marks.keys():
            sheet.cell(row=row,column=1).value = name
            sheet.cell(row=row,column=1).font=font1
            sheet.merge_cells(f'A{row}:C{row+1}')
            sheet.cell(row=row,column=4).value="Total"
            sheet.cell(row=row,column=4).font=font2
            sheet.merge_cells(f'D{row}:D{row+1}')
            row+=2

            for c in range(1,len(marks[name])+1):
                sheet.cell(row=row,column=c).value = marks[name][c-1]
            sheet.cell(row=row,column=len(marks[name])+1).value = f'=SUM(A{row}:C{row})'
            row+=1
        
        
        wb.save("sample_merge.xlsx")
        #cell='=SUM(A1:A3)'
        #freeze pane A2

    
    def sample_bar_chart(self, title,matrix):
        """
            more details
            https://openpyxl.readthedocs.io/en/2.5/charts/chart_layout.html#size-and-position
        """
        rows,cols = len(matrix), len(matrix[0])

        wb = openpyxl.Workbook()
        sheet = wb.active
        
        for data_row in matrix:
            sheet.append(data_row)

        chart = ScatterChart()
        x = Reference(sheet, min_col=1, min_row=1, max_row=rows)
        for i in range(2, cols+1):
            y= Reference(sheet, min_col=i, min_row=1, max_row=rows)
            s =Series(y,x,title_from_data=True)
            chart.series.append(s)
        chart.title=title
        chart.style=12
        chart.x_axis.title="X"
        chart.y_axis.title="Y"
        chart.legend.position='r'

        sheet.add_chart(chart,'B12')
        wb.save("sample_chart.xlsx")